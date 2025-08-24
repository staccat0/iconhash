import time
import hashlib
import urllib3
import base64
import mmh3
import argparse
from openpyxl.styles import Font
import os
import favicon
import requests
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import signal
import sys

# 添加一个全局锁，用于保护datas列表的线程安全
data_lock = threading.Lock()
interrupted = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
SCU = '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"'
xlsx_filename = ""
fingers = []

datas = []
lines = []


def signal_handler(sig, frame):
    global interrupted
    print("\n收到中断信号，正在保存已获取的数据...")
    interrupted = True
    save_data_to_excel()
    sys.exit(0)


def save_data_to_excel():
    print("正在保存已获取的数据到Excel...")
    for data in datas:
        if data[1] is not None:
            save_xlsx(data[0], data[1], data[2], data[3], data[4])
        else:
            workbook = load_workbook(xlsx_filename)
            sheet = workbook.active
            sheet.append(data)
            workbook.save(xlsx_filename)
    print("数据保存完成")


def create():
    try:
        os.makedirs("./icons", exist_ok=True)
    except OSError as e:
        print(f"创建图标目录失败: {e}")

    # 创建xlsx文件
    try:
        wb = Workbook()
        sheet = wb.active
        header_data = ["url", "icon", "hunter", "fofa", "type", "finger", "ico_url"]

        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 50
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 8
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 80
        sheet.append(header_data)
        bold_font = Font(bold=True)
        for cell in sheet[1]:  # sheet[1]表示第一行
            cell.font = bold_font
        global xlsx_filename
        xlsx_filename = f"icons_{time.time()}.xlsx"
        wb.save(xlsx_filename)
        print(f"成功创建Excel文件: {xlsx_filename}")
    except Exception as e:
        print(f"创建Excel文件失败: {e}")


def get_fingers():
    file_path = "fingers.json"
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        # 读取文件内容
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read().strip()
            # 检查文件是否为空
            if not content:
                raise ValueError("文件内容为空")

            # 解析JSON内容
            global fingers
            fingers = json.loads(content)
            return True

    except json.JSONDecodeError as e:
        print(f"JSON解析失败: {str(e)}")
        return False
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        return False


def isfinger(hash):
    hash = str(hash)
    for finger in fingers:
        if hash == finger['hash']:
            return finger['finger']
    return None


def get_icon(url):
    try:
        headers = {'User-Agent': UA, 'referer': url, 'sec-ch-ua': SCU, 'Accept': '*/*'}
        icon = favicon.get(url, headers=headers, timeout=5, verify=False)
        print(f"获取 {url} 的图标成功")
        return icon
    except Exception as e:
        print(f"获取 {url} 的图标失败: {e}")
        return None


def save_icon(icon):
    if not icon:
        print("传入的icon对象为空")
        return False
    try:
        headers = {'User-Agent': UA, 'referer': icon.url, 'sec-ch-ua': SCU, 'Accept': '*/*'}
        response = requests.get(icon.url, headers=headers, verify=False, timeout=5)

        if response.status_code != 200:
            print(f"下载图标失败，状态码: {response.status_code}")
            return False

        iconname = time.time()
        save_path = f"{iconname}.{icon.format}"

        try:
            # img = PILImage.open(BytesIO(response.content))
            # img.save(save_path)
            with open("./icons/" + save_path, 'wb') as f:
                f.write(response.content)
            print(f"图标保存成功: ./icons/{save_path}")
            return save_path
        except Exception as img_e:
            print("图标保存失败")
            return False

    except requests.RequestException as req_e:
        print("图标保存失败")
        return False
    except Exception as e:
        print("图标保存失败")
        return False


def save_xlsx(url, icon_path, icon_format, ico_url, filepath):
    icon_path_tmp = "./icons/" + icon_path
    if not icon_path_tmp or not os.path.exists(icon_path_tmp):
        print(f"图标文件不存在: {icon_path_tmp}")
        return False
    try:
        hunter = get_image_md5(icon_path_tmp)
        fofa = get_image_fofa(icon_path_tmp)
        finger = isfinger(fofa)
        hunter_hash = f'web.icon = "{hunter}"'
        fofa_hash = f'icon_hash="{fofa}"'

        workbook = load_workbook(filepath)
        sheet = workbook.active
        data = [url, None, hunter_hash, fofa_hash, icon_format, finger, ico_url]
        sheet.append(data)
        lines = sheet.max_row

        # 加载图片
        try:
            with open("./icons/" + icon_path, 'rb') as f:
                img = PILImage.open(BytesIO(f.read()))
                img.save("./icons/tmp_" + icon_path)
            img = Image("./icons/tmp_" + icon_path)
            # 将图片添加到指定位置
            sheet.add_image(img, f"B{lines}")
            sheet.row_dimensions[lines].height = 20  # 增加行高以更好显示图标
        except Exception as img_e:
            print(f"添加图片到Excel失败: {img_e}")

        workbook.save(filepath)
        print(f"成功将数据保存到Excel: {url}")
        return ico_url

    except Exception as e:
        print(f"保存数据到Excel失败: {e}")
        return False


def get_image_md5(file_path):
    try:
        if not os.path.exists(file_path):
            print(f"MD5计算失败: 文件 {file_path} 不存在")
            return "N/A"

        md5_hash = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                md5_hash.update(chunk)
        hunter_hash = md5_hash.hexdigest()
        return hunter_hash

    except Exception as e:
        print(f"计算MD5哈希失败: {e}")
        return "N/A"


def get_image_fofa(file_path):
    try:
        if not os.path.exists(file_path):
            print(f"FOFA哈希计算失败: 文件 {file_path} 不存在")
            return "N/A"

        with open(file_path, "rb") as f:
            image_base64 = base64.encodebytes(f.read())
            fofa_hash = mmh3.hash(image_base64)
        return fofa_hash

    except Exception as e:
        print(f"计算FOFA哈希失败: {e}")
        return "N/A"


def input_url(url):
    global interrupted
    if interrupted:
        return

    try:
        print(f"开始处理URL: {url}")
        icon = get_icon(url)
        if not icon:
            print(f"未找到 {url} 的图标")
            # 即使没有图标，也记录URL到Excel
            try:
                with data_lock:
                    data = [url, None, "N/A", "N/A", "N/A", None, "N/A"]
                    datas.append(data)
            except Exception as e:
                print(f"记录无图标的URL失败: {e}")
            return

        # 单线程处理一个URL的所有图标
        for i in icon:
            if interrupted:  # 检查是否收到中断信号
                return
            img_path = save_icon(i)
            if img_path:
                with data_lock:
                    data = (url, img_path, i.format, i.url, xlsx_filename)
                    datas.append(data)
            else:
                print(f"保存图标失败，跳过: {i.url}")

    except Exception as e:
        print(f"处理URL {url} 时发生未知错误: {e}")


def read_file(file_path):
    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"错误: 文件 '{file_path}' 不存在")
            return lines

        # 检查是否为文件
        if not os.path.isfile(file_path):
            print(f"错误: '{file_path}' 不是一个文件")
            return lines

        # 打开文件并按行读取
        with open(file_path, 'r', encoding='utf-8') as file:
            for line_num, line in enumerate(file, 1):
                # 移除行尾的换行符
                cleaned_line = line.rstrip('\n\r')
                lines.append(cleaned_line)
        print(f"\n成功读取文件，共 {len(lines)} 行")
        return lines
    except UnicodeDecodeError:
        print("文件打开失败")


signal.signal(signal.SIGINT, signal_handler)


def run(targer):
    if get_fingers() is True:
        create()
        urls = read_file(targer)
        try:
            # 使用线程池并行处理不同的URL
            with ThreadPoolExecutor(max_workers=10) as executor:
                # 提交所有URL处理任务
                future_to_url = {executor.submit(input_url, url): url for url in urls}

                # 等待所有任务完成
                for future in as_completed(future_to_url):
                    url = future_to_url[future]
                    try:
                        future.result()  # 获取结果，如果有异常会在这里抛出
                    except Exception as e:
                        print(f"处理URL {url} 时发生错误: {e}")
        except KeyboardInterrupt:
            # 处理键盘中断
            print("\n程序被用户中断")
            interrupted = True
        except Exception as e:
            print(f"程序执行过程中发生错误: {e}")
        finally:
            # 无论是否发生异常，都保存数据
            save_data_to_excel()
    else:
        print("finger.json文件不存在")

def main():
    parser = argparse.ArgumentParser()
    # 添加-t参数
    parser.add_argument('-t', type=str, required=True,
                        help='目标文本，文本内格式需带有http://或https://')
    # 解析命令行参数
    args = parser.parse_args()
    # 输出文本
    run(args.t)

if __name__ == "__main__":
    main()
